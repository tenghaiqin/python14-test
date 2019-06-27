#-*- coding: utf-8 -*-
# @Time    : 2019/6/26 18:33
# @Author  : lemon_shan
# @Email   :1529755032@qq.com
# @File    : read_config.py

from configparser import ConfigParser

from openpyxl import load_workbook

from API_1.common import path


class ReadConfig:

    def __init__(self,file_name):
        self.cf =ConfigParser()
        self.cf.read(file_name,encoding='utf-8')

    def get_int(self,section,option):
        value =self.cf.getint(section,option)
        return value

    def get_str(self,section,option):
        value = self.cf.get(section,option)
        return value

    def get_float(self,section,option):
        value = self.cf.getfloat(section,option)
        return value

    def get_bool(self,section,option):
        value = self.cf.getboolean(section,option)
        return value

    def get_data(self,section,option):
        value = self.cf.get(section,option)
        return eval(value)

if __name__ == '__main__':
    res = ReadConfig(path.conf_path).get_str('HOST','host')
    print(res)
    print(type(res))

    wb = load_workbook(path.case_path)
    sheet = wb['register']

    r =res +sheet.cell(2,5).value
    print(r)


