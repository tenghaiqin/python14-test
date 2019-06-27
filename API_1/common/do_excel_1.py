#-*- coding: utf-8 -*-
# @Time    : 2019/6/26 18:05
# @Author  : lemon_shan
# @Email   :1529755032@qq.com
# @File    : do_excel.py
from API_1.common import path
from API_1.common.log import Log
from openpyxl import load_workbook
from API_1.common.read_config import ReadConfig

log = Log()
host = ReadConfig(path.conf_path).get_str('HOST','host')


class Do_excel:

    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def read_data(self,section):
        try:
            wb = load_workbook(self.file_name)
            sheet = wb[self.sheet_name]
            Case_id = ReadConfig(path.conf_path).get_str(section, 'Case_id')
            host = ReadConfig(path.conf_path).get_str('HOST', 'host')
        except Exception as e :
            log.error('打开测试用例出错{}'.format(e))


        test_data = []
        for i in range(2,sheet.max_row+1):
            row_data ={}
            row_data['CaseId']=sheet.cell(i,1).value
            row_data['Module'] = sheet.cell(i, 2).value
            row_data['Title'] = sheet.cell(i, 3).value
            row_data['Method'] = sheet.cell(i, 4).value
            row_data['Url']= host + sheet.cell(i,5).value
            row_data['Param'] = sheet.cell(i, 6).value
            row_data['Sql'] = sheet.cell(i, 7).value
            row_data['ExpectedResult'] = sheet.cell(i, 8).value
            test_data.append(row_data)
        wb.close()
        final_data = []  # 空列表 ，存储最终的测试数据
        if Case_id == 'all':  # CaseId=='all' 获取所有的测试用例
            final_data = test_data  # 把测试数据赋值给final_data
        else:  # 否则如果是列表那就获取列表里面指定id的用例的数据，进行遍历处理
            for i in eval(Case_id):  # 遍历CaseId的值 [1,2,3],获取第1,2,3条的测试用例数据
                final_data.append(test_data[i - 1])  # 实际id值-1，就等于测试数据字典类型的索引值
        return final_data


    def wirte_back(self,row,col,value):

        try:
            wb = load_workbook(self.file_name)
            sheet = wb[self.sheet_name]
        except Exception as e :
            log.error('打开测试用例出错{}'.format(e))
        else:
            sheet.cell(row,col).value = value
            wb.save(self.file_name)
            wb.close()


if __name__ == '__main__':
    file_name=path.case_path
    sheet_name='getlist'

    testdata = Do_excel(file_name,sheet_name).read_data('GetListCase')
    print(testdata)

    # wirte = Do_excel(file_name,sheet_name).wirte_back(row=9,col=1,value='你好')

