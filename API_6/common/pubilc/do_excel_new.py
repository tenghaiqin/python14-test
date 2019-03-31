#-*- coding: utf-8 -*-
#@time    : 2019/3/9 14:46
#@author  : lenmon_shan
#@Email   :1529755032@qq.com
#@file    : Do_Excel.py


from openpyxl import load_workbook
from API_4.common.pubilc import project_path
from API_4.common.pubilc.read_config import ReadConfig
from API_4.common.pubilc.my_log import MyLog
my_log=MyLog()
#拓展
#设置一个思路：一次性读取数据，取哪用例 ，使用配置文件来配置
#配置的数据，是一个字典形式 字典key==表单名

class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name   #excel工作薄的文件或地址
        self.sheet_name = sheet_name  # 表单名


    def read_data(self):
        '''读取excel中所有的测试数据，有返回值'''
        try:
            wb=load_workbook(self.file_name)
            sheet=wb[self.sheet_name]
            Case_id=ReadConfig(project_path.conf_path).get_data('CASE','Case_id')
        except Exception as e:
            my_log.error('文件读取出错了{}'.format(e))

        tel=self.get_tel()#传递表单进来 ，获取一个变量，获取到的值是整数，后面要替换成字符串


        #读取数据开始
        test_data=[]#空列表 ，存储最终的测试数据
        final_data = []  # 空列表 ，存储最终的测试数据
        for key in Case_id:#遍历配置文件的Case_id的值
            sheet_name=key#配置文件里面的key就是表单名
            sheet=wb[sheet_name]#获取一个表单名

            for i in range(2,sheet.max_row+1):
                row_data = {}  # 每一行数据存在一个字典里面
                row_data['CaseId']=sheet.cell(i,1).value
                row_data['Module']=sheet.cell(i,2).value
                row_data['Title']=sheet.cell(i,3).value
                row_data['Method'] = sheet.cell(i, 4).value
                row_data['Url']=sheet.cell(i,5).value
                if sheet.cell(i,6).value.find('tel')!=-1:#如果参数里面有mobile字符串就进行替换
                    row_data['Param'] =sheet.cell(i,6).value.replace('tel',str(tel))#替换函数要字符串类型，所以要加str
                    self.update_tel(int(tel)+1)#更新表格里面每次+1
                else:#如果值不存在，就不做任何替换
                    row_data['Param'] = sheet.cell(i, 6).value
                row_data['ExpectedResult'] = sheet.cell(i, 7).value
                test_data.append(row_data)


            if Case_id[key]=='all':#CaseId=='all' 获取所有的测试用例
                final_data=test_data#把测试数据赋值给final_data
            else:#否则如果是列表那就获取列表里面指定id的用例的数据，进行遍历处理
                for i in Case_id[key]:#遍历CaseId的值 [1,2,3],获取第1,2,3条的测试用例数据
                    final_data.append(test_data[i-1])#实际id值-1，就等于测试数据字典类型的索引值
            wb.close()
        return final_data

    def get_tel(self):
        '''获取存在Excel里面的文件'''
        wb = load_workbook(self.file_name)
        sheet = wb['mobile']
        wb.close()
        return sheet.cell(1,2).value#返回电话号码的值

    def update_tel(self,new_tel):
        '''写回手机号码'''
        wb=load_workbook(self.file_name)
        sheet=wb['mobile']#指定表单名也可以灵活变动
        sheet.cell(1,2,new_tel)#指定位置更新手机号码
        wb.save(self.file_name)
        wb.close()

    def write_back(self,row,col,value):
        '''在指定的单元格写入指定的数据，并保存到当前Excel
        row:代表写回的行数
        col:代表入的列数
        value:代表写入的值'''

        try:
            wb=load_workbook(self.file_name)
            sheet=wb[self.sheet_name]
        except Exception as e:
            print('出错{}'.format(e))
        else:
            sheet.cell(row,col).value=value#为excel指定的行列写入指定的值
            wb.save(self.file_name)#写完结果后，要保存
            wb.close()#每次操作完 关闭掉！！！

if __name__ == '__main__':

    file_name='test_api.xlsx'
    sheet_name='recharge'
    test_data=DoExcel(project_path.case_path,sheet_name).read_data()
    # test_data_1=DoExcel('test_api.xlsx','Sheet1').write_back(20,9,'你好')
    # print(DoExcel(project_path.case_path, 'mobile').get_mobile())
    print(test_data)

